import subprocess
from pathlib import Path
from docx import Document
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import load_workbook
import xlrd


class ParseError(RuntimeError):
    pass


def _parse_doc(path: Path) -> str:
    result = subprocess.run(["antiword", str(path)], capture_output=True, timeout=120)
    if result.returncode != 0:
        raise ParseError(f"Antiword parsing failed: {result.stderr.decode(errors='replace')[-500:]}")
    text = result.stdout.decode("utf-8", errors="replace").strip()
    if not text:
        raise ParseError("Word document contains no extractable text")
    return f"[文件] {path.name}\n{text}"


def _parse_xls(path: Path) -> str:
    workbook = xlrd.open_workbook(path, on_demand=True)
    parts: list[str] = [f"[文件] {path.name}"]
    for sheet in workbook.sheets():
        if getattr(sheet, "visibility", 0) != 0:
            continue
        parts.append(f"\n[工作表] {sheet.name}")
        for row_index in range(sheet.nrows):
            values = [str(sheet.cell_value(row_index, column_index)) for column_index in range(sheet.ncols)]
            while values and not values[-1]:
                values.pop()
            if any(values):
                parts.append(" | ".join(values))
    workbook.release_resources()
    if len(parts) <= 1:
        raise ParseError("Excel workbook contains no visible data")
    return "\n".join(parts)


def _parse_docx(path: Path) -> str:
    document = Document(path)
    parts: list[str] = [f"[文件] {path.name}"]
    for element in document.iter_inner_content():
        if isinstance(element, Paragraph):
            text = element.text.strip()
            if text:
                style = element.style.name if element.style else "正文"
                parts.append(f"[{style}] {text}")
        elif isinstance(element, Table):
            parts.append("[表格]")
            for row in element.rows:
                parts.append(" | ".join(cell.text.strip().replace("\n", " ") for cell in row.cells))
    if len(parts) <= 1:
        raise ParseError("Word document contains no extractable text")
    return "\n".join(parts)


def _parse_xlsx(path: Path) -> str:
    formulas = load_workbook(path, read_only=True, data_only=False)
    values = load_workbook(path, read_only=True, data_only=True)
    parts: list[str] = [f"[文件] {path.name}"]
    for sheet in formulas.worksheets:
        if sheet.sheet_state != "visible":
            continue
        value_sheet = values[sheet.title]
        parts.append(f"\n[工作表] {sheet.title}")
        for formula_row, value_row in zip(sheet.iter_rows(), value_sheet.iter_rows()):
            cells: list[str] = []
            for formula_cell, value_cell in zip(formula_row, value_row):
                formula, value = formula_cell.value, value_cell.value
                cells.append(f"{formula} => {value or ''}" if isinstance(formula, str) and formula.startswith("=") else ("" if value is None else str(value)))
            while cells and not cells[-1]:
                cells.pop()
            if any(cells):
                parts.append(" | ".join(cells))
    formulas.close()
    values.close()
    if len(parts) <= 1:
        raise ParseError("Excel workbook contains no visible data")
    return "\n".join(parts)


def extract_document(path_value: str) -> str:
    path = Path(path_value)
    try:
        suffix = path.suffix.lower()
        if suffix == ".doc":
            return _parse_doc(path)
        if suffix == ".xls":
            return _parse_xls(path)
        if suffix == ".docx":
            return _parse_docx(path)
        if suffix == ".xlsx":
            return _parse_xlsx(path)
        raise ParseError(f"Unsupported file extension: {suffix}")
    except ParseError:
        raise
    except Exception as exc:
        raise ParseError(str(exc)) from exc
